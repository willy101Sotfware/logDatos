import json
import pandas as pd
import matplotlib.pyplot as plt
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font

def cargar_datos(file_path):
    """Carga los datos desde un archivo JSON."""
    with open(file_path, 'r', encoding='utf-8') as file:
        return json.load(file)

def procesar_log(file_path):
    """Procesa el log y extrae información relevante."""
    with open(file_path, 'r', encoding='utf-8') as file:
        lineas = file.readlines()

    datos_organizados = []

    for linea in lineas:
        linea = linea.strip()
        if linea:
            # Extraer la fecha y hora
            fecha_hora = re.search(r'(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2})', linea)
            fecha = fecha_hora.group(1) if fecha_hora else None
            
            # Extraer información del mensaje XML
            xml_data = re.search(r'<string>(.*?)</string>', linea)
            if xml_data:
                xml_content = xml_data.group(1)
                
                # Extraer campos específicos del contenido XML
                identification = re.search(r'Identification:\s*([0-9]+)', xml_content)
                valor_aprobado = re.search(r'Valor Aprobado:\s*([0-9]+)', xml_content)
                usuario = re.search(r'Usuario:\s*([A-Z0-9_]+)', xml_content)
                password = re.search(r'Pass:\s*([^\s]+)', xml_content)
                valor_recolectado = re.search(r'Valor recolectado:\s*([0-9]+)', xml_content)
                codigo_trans = re.search(r'Codigo Trans:\s*([0-9]+)', xml_content)
                celular = re.search(r'Celular:\s*([0-9]+)', xml_content)
                email = re.search(r'Email:\s*([^\s]+)', xml_content)
                habeas = re.search(r'Habeas:\s*(True|False)', xml_content)
                recibo = re.search(r'Recibo:\s*(True|False)', xml_content)
                manera_pago = re.search(r'Manera Pago:\s*([A-Z])', xml_content)
                aprove_val = re.search(r'AproveVal:\s*([0-9]+)', xml_content)

                # Agregar los datos organizados a la lista
                datos_organizados.append({
                    'fecha': fecha,
                    'identification': identification.group(1) if identification else None,
                    'valor_aprobado': valor_aprobado.group(1) if valor_aprobado else None,
                    'usuario': usuario.group(1) if usuario else None,
                    'password': password.group(1) if password else None,
                    'valor_recolectado': valor_recolectado.group(1) if valor_recolectado else None,
                    'codigo_trans': codigo_trans.group(1) if codigo_trans else None,
                    'celular': celular.group(1) if celular else None,
                    'email': email.group(1) if email else None,
                    'habeas': habeas.group(1) if habeas else None,
                    'recibo': recibo.group(1) if recibo else None,
                    'manera_pago': manera_pago.group(1) if manera_pago else None,
                    'aprove_val': aprove_val.group(1) if aprove_val else None
                })

    # Guardar los datos organizados en un nuevo archivo JSON
    with open('LogOrganizado.json', 'w', encoding='utf-8') as json_file:
        json.dump(datos_organizados, json_file, ensure_ascii=False, indent=4)

def guardar_excel(df):
    """Guarda los datos analizados en un archivo Excel con formato."""
    # Guardar el DataFrame en un archivo Excel
    df.to_excel('datos_analizados.xlsx', index=False)

    # Abrir el archivo Excel para formatear
    wb = Workbook()
    ws = wb.active

    # Escribir los encabezados
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(color="FF0000")  # Color rojo para los encabezados

    # Escribir los datos
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Guardar el archivo Excel
    wb.save('datos_analizados.xlsx')

def graficar_datos(df):
    """Grafica la frecuencia de los estados."""
    conteo_estados = df['recibo'].value_counts()
    plt.figure(figsize=(8, 5))
    conteo_estados.plot(kind='bar', color='skyblue')
    plt.title('Frecuencia de Recibos')
    plt.xlabel('Recibo')
    plt.ylabel('Frecuencia')
    plt.xticks(rotation=0)
    plt.tight_layout()
    plt.savefig('frecuencia_recibos.png')
    plt.show()

if __name__ == "__main__":
    # Intentar eliminar archivos existentes si existen
    for archivo in ['datos_analizados.xlsx', 'frecuencia_recibos.png', 'LogOrganizado.json']:
        try:
            if os.path.exists(archivo):
                os.remove(archivo)
        except PermissionError:
            print(f"No se pudo eliminar {archivo}. Asegúrate de que no esté abierto en otro programa.")

    try:
        # Procesar el log
        procesar_log('Log2025-01-17.json')
        
        # Cargar los datos organizados
        datos = cargar_datos('LogOrganizado.json')
        df = pd.DataFrame(datos)  # Convertir los datos a un DataFrame
        
        # Guardar en Excel y graficar
        guardar_excel(df)
        graficar_datos(df)
    except FileNotFoundError as e:
        print(f"Error: {e}. Asegúrate de que el archivo exista.")
    except Exception as e:
        print(f"Ocurrió un error: {e}") 